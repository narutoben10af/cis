import openpyxl
wb = openpyxl.load_workbook("Assignment07.xlsx")

sheet = wb["Survey"]
sheet2 = wb["Error Locations"]
wb.create_sheet("FixedSurvey")
sheet3 = wb["FixedSurvey"]

nMaxRow = sheet.max_row
nMaxCol = sheet.max_column

print("Max column:", nMaxCol)
print("Max row:", nMaxRow)

print("Yes")
x = 3
y = 3

for i in range(1, nMaxRow + 1):
    for j in range(1, nMaxCol+1):
        sheet3.cell(row = i, column = j).value = sheet.cell(row = i, column = j).value
        strValue = sheet3.cell(row = i, column = j).value
        strLocation = sheet3.cell(row = i, column = j)
        if (strValue < -3):
             sheet2.cell(row = x, column = 1).value = strLocation.coordinate
             sheet2.cell(row = x, column = 2).value = strValue
             x += 1
             strValue = -3
             sheet3.cell(row=i, column=j).value = strValue
        if (strValue > 3):
            sheet2.cell(row = y, column=4).value = strLocation.coordinate
            sheet2.cell(row=y, column=5).value = strValue
            y+= 1
            strValue = 3
            sheet3.cell(row=i, column=j).value = strValue

print("Yes")

wb.save("Assignment07 - copy2.xlsx")
print("Yes")

