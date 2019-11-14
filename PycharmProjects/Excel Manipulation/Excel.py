import openpyxl
from openpyxl.utils import get_column_letter

strColumn = get_column_letter(1)
print(strColumn)

strColumn = get_column_letter(10)
print(strColumn)

strColumn = get_column_letter(27)
print(strColumn)

strColumn = get_column_letter(8)
print(strColumn, end = "")

strColumn = get_column_letter(1)
print(strColumn, end = "")

strColumn = get_column_letter(14)
print(strColumn, end = "")

strColumn = get_column_letter(9)
print(strColumn)


from openpyxl.utils import column_index_from_string

nColumn = column_index_from_string("H")
print(nColumn)

nColumn = column_index_from_string("A")
print(nColumn)

nColumn = column_index_from_string("N")
print(nColumn)

nColumn = column_index_from_string("I")
print(nColumn)

wb = openpyxl.load_workbook("example.xlsx")

#Get the sheets in the workbook - old style
# listSheets = wb.get_sheet_names()
# print(listSheets)

#Get the sheets in the workbook - recent style
listSheets = wb.sheetnames
print(listSheets)

#Get a sheet by name - recent style
sheet = wb["Sheet1"]
print(sheet.title)

strValue = sheet["A1"].value
print(strValue)

cell = sheet.cell(row = 1, column = 2)
# cell = sheet["B1"]
print("Row is", str(cell.row)) #Row 1
print("Column is", str(cell.column)) #B
print("Value is", str(cell.value))
print("Cell", cell.coordinate, "is", cell.value) #B1

for i in range(1,8):
    print(i, sheet.cell(row = i, column = 2).value)

# for i in range(1, 8):
#     for j in range (1, 4):
#       print(i, sheet.cell(row = i, column = j).value)

for i in range(1,10):
    print(i, sheet.cell(row = i, column = 2).value)

nMaxRow = sheet.max_row
nMaxCol = sheet.max_column

print("Max row:", nMaxRow)
print("Max column:", nMaxCol)

for i in range(1, nMaxRow + 1):
    for j in range (1, nMaxCol + 1):
      print(i, sheet.cell(row = i, column = j).value, end = "\t\t\t")

    print()


