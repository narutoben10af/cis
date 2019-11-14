import openpyxl
wb = openpyxl.Workbook() #create new blank workbook

sheet = wb.active #Create an empty workbook

sheet.title = "Attendance" #change the title of the sheet

wb.save("ExcelPython.xlsx") # save and pass name as parameter
print("Done")


#Creates a copy of the workbook
wb = openpyxl.load_workbook("ExcelPython.xlsx")

sheet = wb["Attendance"]

sheet.title = "New Attendance"  # Change the title of the sheet

wb.save("ExcelPython - copy.xlsx") #save and pass name as parameter
print("Done")

#Add or remove worksheets

wb = openpyxl.load_workbook("ExcelPython - copy.xlsx")

wb.create_sheet("Fruits") #Creates a new sheet
wb.create_sheet("Vegetables")

listSheets = wb.sheetnames
print(listSheets)

wb.remove(wb["New Attendance"]) # Removes a sheet by name
listSheets = wb.sheetnames
print(listSheets)

wb.save("ExcelPython - copy.xlsx") #Save current workbook
print("Done")

wb = openpyxl.load_workbook("ExcelPython - copy.xlsx")

sheet = wb["Fruits"]
sheet["A1"] = "Apple"
sheet["A2"] = "Banana"
sheet["A3"] = "Cherries"

wb.save("ExcelPython - copy.xlsx") #save current workbook
print("Done")

#Exercise
import openpyxl
wb = openpyxl.load_workbook("produceSales.xlsx")
sheet = wb["Sheet"]
nMaxRow = sheet.max_row
for i in range(2, nMaxRow + 1):
    strValue = sheet.cell(row = i, column = 1)
    strValue2 = sheet.cell(row = i, column = 2)
    if(strValue == "Garlic"):
        strValue2 = 3.07
    if(strValue == "Celery"):
        strValue2 = 1.19
    if(strValue == "Lemon"):
        strValue2 = 1.27
    #sheet.cell(row = i, column = 2).value = strValue2


wb.save("produceSales - copy.xlsx")

print("Done")

#Solution

wb = openpyxl.load_workbook("produceSales.xlsx")
sheet = wb["Sheet"]

#The produce types and their update prices
PRICE_UPDATES = {"Garlic": 3.07, "Celery" : 1.19, "Lemon" : 1.27}

#Loop through the rows and update the prices
for rowNum in range(2, sheet.max_row):
    produceName = sheet.cell(row = rowNum, column = 1).value
    if produceName in PRICE_UPDATES:
        sheet.cell(row = rowNum, column = 2).value = PRICE_UPDATES

wb.save("updatedProduceSales.xlsx")
print("Done")